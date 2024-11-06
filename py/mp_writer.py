import openpyxl
import pandas
import xlsxwriter
import shutil

from marking_plan import MarkingPlan
from html_reader import HTMLReader


class MPWriter:
    def __init__(self):
        self.__marking_plan = None

        self.__filename = None

    def __write(self, fname, url):
        self.__filename = fname

        shutil.copyfile("xls/_.xlsx", self.__filename)
		
        sheets = openpyxl.load_workbook(self.__filename)
        sheet = sheets['Spécifications techniques']

        reader = HTMLReader()
        reader.read_url(url)

        page = reader.render()

        self.__marking_plan.put(("page_view", True, "Sur tout le site",
                                 "Au chargement de la page", "Collecté automatiquement"))

        subscribe = page.get_attr_tags('value', ['Mon compte', 'identifi'], False)

        if not subscribe:
            subscribe = page.get_data_tags('Mon compte')

        if len(subscribe) > 0 and (subscribe[0].get_name() == 'button' or subscribe[0].get_name() == 'a'):
            subs = HTMLReader()
            subs.read_url(url)
            self.__marking_plan.put(("subscribe", True, "Lorsque l'utilisateur crée un compte",
                                     "Au chargement de la page d'inscription", "Recommandé"))

        login = page.get_attr_tags('value', ['connect', 'login'], False)

        if not login:
            login = page.get_data_tags('Se connecter')

            if not login:
                login = page.get_data_tags('Mon compte')

        if len(login) > 0 and (login[0].get_name() == 'button' or login[0].get_name() == 'a'):
            self.__marking_plan.put(("login", True, "Lorsque l'utilisateur se connecte",
                                     "Au clic sur 'se connecter", "Recommandé"))

        search = page.get_attr_tags('value', ['Recherche', 'rechercher'], False)

        if len(search) > 0 and (search[0].get_name() == 'input' or search[0].get_name() == 'button'):
            self.__marking_plan.put(("search", True, "Lorsque l'utilisateur effectue une recherche",
                                     "Au chargement de la page de search", "Recommandé"))

        share = page.get_attr_tags('value', ['Partage', 'share'], False)

        if not share:
            share = page.get_data_tags('Suivez-nous')

        if len(share) > 0 and (share[0].get_name() == 'input' or share[0].get_name() == 'button'
                               or share[0].get_name() == 'a'):
            self.__marking_plan.put(("share", True, "Lorsque l'utilisateur partage du contenu",
                                     "Au clic sur les boutons de partage du contenu", "Recommandé"))

        news = page.get_attr_tags('name', ['email', 'news'], False)

        if len(news) > 0 and (news[0].get_name() == 'input' or news[0].get_name() == 'button'):
            self.__marking_plan.put(("newsletter_subscription", True, "Lorsque l'utilisateur s'inscrit à la newsletter",
                                     "Au clic de s'inscrire à la newsletter", "Personnalisé"))

        search_channel = page.get_attr_tags('class', 'travelcheck')

        if search_channel:
            self.__marking_plan.put(("search_channel", True, "Lorsque l'utilisateur sélectionne un canal de recherche",
                                     "Au clic sur le toggle des canaux de recherche", "Personnalisé"))

        tel = page.get_attr_tags('href', ['tel'], False)

        if tel:
            self.__marking_plan.put(("telephone", True, "Lorsque l'utilisateur clique sur le bouton 'téléphoner'",
                                     "Au clic sur le numéro de téléphone", "Personnalisé"))

        contact = page.get_attr_tags('href', ['contact'], False)

        if not contact:
            contact = page.get_data_tags('Contactez-nous')

            if not contact:
                print('r')
                contact = page.get_data_tags('Nous contacter')

        if contact:
            self.__marking_plan.put(("contact", True, "Lorsque l'utilisateur clique sur le bouton 'Contacter'",
                                     "Au clic sur le bouton de contact", "Personnalisé"))

        items = page.get_attr_tags('class', ['item'], False)

        if not items:
            items = page.get_data_tags('item')

        if items:
            self.__marking_plan.put(("view_item", True, "A la consultation de la page produit",
                                     "Au chargement de page", "Recommandé"))

            self.__marking_plan.put(("add_to_wishlist", False, "Si un utilisateur fait un ajout",
                                     "Au clic sur le contenu", "Recommandé"))

            self.__marking_plan.put(("select_item", True, "Lors de la sélection d'un item dans une liste",
                                     "Au clic sur le produit sélectionné", "Recommandé"))

        self.__marking_plan.put(("select_content", True, "Lorsque l'utilisateur sélectionne un contenu",
                                 "Au clic sur le contenu", "Recommandé"))

    def build(self, fname, url):
        self.__marking_plan = MarkingPlan()

        self.__write(fname, url)

    def render(self):
        sheets = openpyxl.load_workbook(self.__filename)
        sheet = sheets['Spécifications techniques']

        for i in range(self.__marking_plan.count()):
            entry = self.__marking_plan.get_index(i)

            sheet['B' + str(i + 12)] = '✔' if entry[1] else '✗'
            sheet['C' + str(i + 12)] = entry[2]
            sheet['D' + str(i + 12)] = entry[3]
            sheet['E' + str(i + 12)] = entry[0]
            sheet['F' + str(i + 12)] = entry[4]

            i += 1

        for i in range(self.__marking_plan.count()):
            sheet['H' + str(i + 12)] = 'NOK'
            sheet['I' + str(i + 12)] = 'NOK'
            sheet['J' + str(i + 12)] = 'NOK'
            sheet['K' + str(i + 12)] = 'NOK'

        sheets.save(self.__filename)

    def write_dl(self, dl):
        sheets = openpyxl.load_workbook(self.__filename)
        sheet = sheets['Spécifications techniques']

        for i in range(self.__marking_plan.count()):
            sheet['G' + str(i + 12)] = dl.render(i)

        sheets.save(self.__filename)
